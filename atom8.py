import json
import sys
import os
import logging
import time
from datetime import datetime
from PyQt5.QtCore import Qt, QSize, QRect
from PyQt5.QtGui import QColor, QTextFormat, QPainter, QPixmap, QIcon
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QLineEdit, QLabel, QComboBox, \
    QListWidget, QHBoxLayout, QAction, QMessageBox, QFileDialog, QStatusBar, QCheckBox, QTextEdit, QInputDialog, \
    QDialog, QTableWidgetItem, QTableWidget, QMenu, QHeaderView, QPlainTextEdit, QTabWidget, QGroupBox, QScrollArea, \
    QSplashScreen, QMenuBar, QFrame
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.edge.options import Options as EdgeOptions
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from helper import extract_elements_to_json
import platform
import cv2

__version__ = "0.0.3"
__build__ = f"{datetime.now().strftime('%Y%m%d')}"


class CustomComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self.showContextMenu)

    def showContextMenu(self, position):
        try:
            menu = QMenu()
            copyAction = menu.addAction("Copy")
            action = menu.exec_(self.mapToGlobal(position))
            if action == copyAction:
                self.copyToClipboard()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while copying to clipboard: {e}")

    def copyToClipboard(self):
        try:
            text = self.currentText()

            parts = text.split(":", 1)
            if len(parts) == 2:
                text_to_copy = parts[1].strip()
            else:
                text_to_copy = text

            clipboard = QApplication.clipboard()
            clipboard.setText(text_to_copy)
            self.logger.info(f"Copied {text_to_copy} to clipboard.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while copying to clipboard: {e}")


class QTextEditLogger(logging.Handler):
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.widget.setReadOnly(True)

    def emit(self, record):
        msg = self.format(record)
        self.widget.append(msg)


class LineNumberArea(QWidget):
    def __init__(self, editor):
        super().__init__(editor)
        self.codeEditor = editor

    def sizeHint(self):
        return QSize(self.codeEditor.lineNumberAreaWidth(), 0)

    def paintEvent(self, event):
        self.codeEditor.lineNumberAreaPaintEvent(event)


class ScriptEditor(QPlainTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.lineNumberArea = LineNumberArea(self)

        self.blockCountChanged.connect(self.updateLineNumberAreaWidth)
        self.updateRequest.connect(self.updateLineNumberArea)
        self.cursorPositionChanged.connect(self.highlightCurrentLine)
        self.updateLineNumberAreaWidth(0)

    def lineNumberAreaWidth(self):
        digits = 1
        max_count = max(1, self.blockCount())
        while max_count >= 10:
            max_count /= 10
            digits += 1
        space = 3 + self.fontMetrics().width('9') * digits
        return space

    def updateLineNumberAreaWidth(self, _):
        self.setViewportMargins(self.lineNumberAreaWidth(), 0, 0, 0)

    def updateLineNumberArea(self, rect, dy):
        if dy:
            self.lineNumberArea.scroll(0, dy)
        else:
            self.lineNumberArea.update(0, rect.y(), self.lineNumberArea.width(), rect.height())

        if rect.contains(self.viewport().rect()):
            self.updateLineNumberAreaWidth(0)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        cr = self.contentsRect()
        self.lineNumberArea.setGeometry(QRect(cr.left(), cr.top(), self.lineNumberAreaWidth(), cr.height()))

    def lineNumberAreaPaintEvent(self, event):
        painter = QPainter(self.lineNumberArea)
        painter.fillRect(event.rect(), Qt.lightGray)

        block = self.firstVisibleBlock()
        block_number = block.blockNumber()
        top = self.blockBoundingGeometry(block).translated(self.contentOffset()).top()
        bottom = top + self.blockBoundingRect(block).height()

        while block.isValid() and top <= event.rect().bottom():
            if block.isVisible() and bottom >= event.rect().top():
                number = str(block_number + 1)
                painter.setPen(Qt.black)
                painter.drawText(0, top, self.lineNumberArea.width(), self.fontMetrics().height(),
                                 Qt.AlignRight, number)

            block = block.next()
            top = bottom
            bottom = top + self.blockBoundingRect(block).height()
            block_number += 1

    def highlightCurrentLine(self):
        extraSelections = []

        if not self.isReadOnly():
            selection = QTextEdit.ExtraSelection()
            lineColor = QColor(Qt.yellow).lighter(160)
            selection.format.setBackground(lineColor)
            selection.format.setProperty(QTextFormat.FullWidthSelection, True)
            selection.cursor = self.textCursor()
            selection.cursor.clearSelection()
            extraSelections.append(selection)

        self.setExtraSelections(extraSelections)


class Atom8(QMainWindow):
    def __init__(self):
        super().__init__()
        self.splash = QSplashScreen(QPixmap("_internal/assets/splash.png"))
        splash_message = f"Loading Atom8...\nVersion: v{__version__} build {__build__}-dev"
        self.splash.showMessage(splash_message, Qt.AlignLeft | Qt.AlignBottom, Qt.white)

        self.splash.show()
        time.sleep(3)
        self.splash.finish(self)

        self.setWindowIcon(QIcon("_internal/assets/atom-8-icon.png"))

        self.driver = None
        self.steps = []
        self.recentFiles = []
        self.recentFilesMenu = None
        self.initUI()
        self.setupLogging()
        self.loadRecentFiles()
        self.currentFilePath = None
        self.resultsTable = None
        self.setupScriptEditor()
        self.results = []
        self.loadWhenDone = False

    def initUI(self):

        style = """
        QWidget {
            background-color: #FFFFFF;
            font-size: 12px;
        }

        QPushButton {
            color: white;
            background-color: #007BFF;
            border-radius: 4px;
            padding: 6px 12px;
            border: none;
            font-size: 12px;
        }

        QPushButton:hover {
            background-color: #0069D9;
            border-color: #0062CC;
        }

        QPushButton:pressed {
            background-color: #005CBF;
            border-color: #0056B3;
        }

        QLabel {
            color: #555;
        }

        QLineEdit {
            color: #555;
            border: 1px solid #ddd;
            padding: 6px;
            border-radius: 4px;
            background-color: #fafafa;
        }
        
        QLineEdit:diabled {
            background-color: #eee;
        }

        QLineEdit:focus {
            border-color: #007BFF;
            outline: none;
        }

        QStatusBar {
            background-color: #F7F7F7;
            color: #555;
        }

        QComboBox {
            border: 1px solid #ddd;
            padding: 10px;
            border-radius: 4px;
            background-color: #fff;
            color: #555;
        }

        QComboBox::drop-down {
            background-color: transparent;
        }

        QComboBox::down-arrow {
            image: url(_internal/assets/drop-down-arrow.png);
            padding-right: 20px;
            width: 10px;
            height: 10px;
        }

        QComboBox QAbstractItemView {
            background-color: #fff;
            color: #555;
        }

        QListWidget {
            border: 1px solid #ddd;
            border-radius: 4px;
            color: #555;
            background-color: #f5f5f5;
        }

        QListWidget::item {
            padding: 4px;
            color: #555;
        }

        QListWidget::item:selected {
            background-color: #007BFF;
            color: white;
        }

        QTextEdit {
            border: 1px solid #ddd;
            color: #333;
            background-color: #f5f5f5;
        }

        QCheckBox {
            color: #555;
        }

        QMenuBar {
            color: #333;
        }

        QMenuBar::item {
            background-color: transparent;
        }

        QMenuBar::item:selected { 
            background-color: #D6D6D6;
        }

        QMenuBar::item:pressed {
            background-color: #C6C6C6;
        }

        QMenu {
            background-color: #FFFFFF;
            border: 1px solid #ddd;
        }

        QMenu::item {
            padding: 6px;
            width: 150px;
        }

        QMenu::item:selected {
            background-color: #007BFF;
            color: white;
        }

        QMenu::item:pressed {
            background-color: #0069D9;
            color: white;
        }

        QGroupBox {
            background-color: #FFFFFF;
            border: 1px solid #ddd;
            border-radius: 4px;
            padding: 10px;
        }

        QScrollArea {
            border: 1px solid #ddd;
            border-radius: 4px;
        }
        QScrollBar:vertical {
            border: none;
            background: #EEEEEE;
            width: 10px;
            margin: 10px 0 10px 0;
        }
        QScrollBar::handle:vertical {
            background: #d1d1d1;
            min-height: 20px;
            border-radius: 4px;
        }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            border: none;
            background: none;
        }
        QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {
            border: none;
            background: none;
        }
        QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
            background: none;
        }
        
        QScrollBar:horizontal {
            border: none;
            background: #EEEEEE;
            height: 10px;
            margin: 0 10px 0 10px;
        }
        
        QScrollBar::handle:horizontal {
            background: #d1d1d1;
            min-width: 20px;
            border-radius: 4px;
        }
        
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
            border: none;
            background: none;
        }
        
        QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {
            border: none;
            background: none;
        }
        
        QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
            background: none;
        }

        QPushButton.secondary-btn {
            color: #333;
            background-color: #FFFFFF;
            border-radius: 4px;
            padding: 6px 12px;
            border: 1px solid #ddd;
            font-size: 12px;
        }

        QPushButton.secondary-btn:hover {
            background-color: #EEEEEE;
            border-color: #ddd;
            cursor: pointer;
        }
        """
        self.setWindowTitle(f"Atom8 v{__version__}")
        self.setGeometry(100, 100, 800, 800)
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

        self.testName = QLineEdit(self)
        self.testName.setPlaceholderText("Test Name")

        self.testDescription = QLineEdit(self)
        self.testDescription.setPlaceholderText("Description")

        mainLayout = QVBoxLayout()
        mainLayout.addWidget(self.testName)
        mainLayout.addWidget(self.testDescription)

        self.setupMenuBar()
        self.setupActionSelection(mainLayout)
        self.setupButtonsAndStepsList(mainLayout)

        mainTabWidget = QTabWidget(self)
        mainLayout.addWidget(mainTabWidget)

        logViewerTab = QWidget()
        logViewerLayout = QVBoxLayout(logViewerTab)
        self.logViewer = QTextEdit(self)
        self.logViewer.setReadOnly(True)
        logViewerLayout.addWidget(self.logViewer)

        buttonsLayout = QHBoxLayout()
        self.saveLogsButton = QPushButton('Save Logs', self)
        self.saveLogsButton.clicked.connect(self.saveLogs)
        buttonsLayout.addWidget(self.saveLogsButton)
        self.clearLogsButton = QPushButton('Clear Logs', self)
        self.clearLogsButton.clicked.connect(self.clearLogs)
        self.clearLogsButton.setProperty("class", "secondary-btn")
        buttonsLayout.addWidget(self.clearLogsButton)
        logViewerLayout.addLayout(buttonsLayout)
        mainTabWidget.addTab(logViewerTab, "Log Viewer")

        browserOptionsTab = QWidget()
        browserOptionsLayout = QVBoxLayout(browserOptionsTab)
        mainTabWidget.addTab(browserOptionsTab, "Browser Options")

        basicOptionsScrollArea = QScrollArea()
        basicOptionsScrollArea.setWidgetResizable(True)
        basicOptionsScrollArea.setMaximumHeight(130)
        basicOptionsGroup = QGroupBox("Basic Options")
        basicOptionsLayout = QVBoxLayout(basicOptionsGroup)
        basicOptionsScrollArea.setWidget(basicOptionsGroup)
        browserOptionsLayout.addWidget(basicOptionsScrollArea)

        basicOptions = [
            ("Headless Mode", "Run the browser in the background without GUI"),
            ("Disable GPU", "Disable GPU acceleration"),
            ("Incognito Mode", "Run browser in incognito/private mode"),
            ("Disable Popup Blocking", "Disable popup blocking feature"),
            ("Disable Infobars", "Disable infobars in the browser"),
            ("Disable Extensions", "Disable extensions in the browser"),
        ]

        for option in basicOptions:
            checkbox = self.createCheckbox(option[0], option[1])
            basicOptionsLayout.addWidget(checkbox)

        advancedOptionsScrollArea = QScrollArea()
        advancedOptionsScrollArea.setWidgetResizable(True)
        advancedOptionsScrollArea.setMaximumHeight(130)
        advancedOptionsGroup = QGroupBox("Advanced Options")
        advancedOptionsLayout = QVBoxLayout(advancedOptionsGroup)
        advancedOptionsScrollArea.setWidget(advancedOptionsGroup)
        browserOptionsLayout.addWidget(advancedOptionsScrollArea)

        advancedOptions = [
            ("Disable Dev Shm Usage", "Disable Dev Shm Usage (Chrome only)"),
            ("Ignore Certificate Errors", "Ignore SSL certificate errors"),
            ("Custom User Agent", "Set a custom user agent string"),
            ("Disable JavaScript", "Disable JavaScript execution in the browser"),
            ("Disable Images", "Disable image loading in the browser"),
            ("Enable Network Throttling", "Simulate different network conditions"),
            ("Enable Performance Logging", "Enable logging of performance metrics"),
            ("Enable GPU Hardware Acceleration", "Enable GPU hardware acceleration"),
            ("Remote Debugging Port", "Set a remote debugging port"),
            ("Proxy Settings", "Configure proxy settings for the browser"),
            ("Enable Automation", "Enable automation flags in the browser"),
            ("No Sandbox", "Disable the sandbox for elevated privileges"),
            ("Disable Web Security", "Disable web security features"),
            ("Enable Experimental Features", "Enable experimental features in the browser"),
            ("Disable Password Manager", "Disable the browser's password manager"),
            ("Disable Autofill", "Disable form autofill in the browser"),
            ("Disable Filesystem API", "Disable the Filesystem API"),
            ("Disable Geolocation", "Disable geolocation features"),
        ]

        for option in advancedOptions:
            checkbox = self.createCheckbox(option[0], option[1])
            advancedOptionsLayout.addWidget(checkbox)

        self.browserLabel = QLabel("Browser: " + self.loadSetting("defaultBrowser", "Chrome"))
        self.browserLabel.setStyleSheet("margin-top: 10px; margin-bottom: 10px;")
        mainLayout.addWidget(self.browserLabel)

        self.setStyleSheet(style)
        centralWidget = QWidget()
        centralWidget.setLayout(mainLayout)
        self.setCentralWidget(centralWidget)

    def setupMenuBar(self):
        try:
            menuBar = self.menuBar()
            fileMenu = menuBar.addMenu('File')
            toolsMenu = menuBar.addMenu('Tools')
            helpMenu = menuBar.addMenu('Help')

            newAction = QAction('New', self)
            newAction.triggered.connect(self.newFile)
            newAction.setShortcut('Ctrl+N')
            fileMenu.addAction(newAction)

            openAction = QAction('Open', self)
            openAction.triggered.connect(self.openFile)
            openAction.setShortcut('Ctrl+O')
            fileMenu.addAction(openAction)

            self.recentFilesMenu = fileMenu.addMenu('Open Recent')
            self.recentFilesMenu.setStyleSheet("QMenu::item { width: 250px; }")

            self.updateRecentFilesMenu()

            realSaveAction = QAction('Save', self)
            realSaveAction.triggered.connect(self.realSaveFile)
            realSaveAction.setShortcut('Ctrl+S')
            fileMenu.addAction(realSaveAction)

            saveAction = QAction('Save As', self)
            saveAction.triggered.connect(self.saveFile)
            saveAction.setShortcut('Ctrl+Shift+S')
            fileMenu.addAction(saveAction)

            clearAction = QAction('Clear All', self)
            clearAction.triggered.connect(self.clearAllFields)
            clearAction.setShortcut('Ctrl+N')
            fileMenu.addAction(clearAction)

            fileMenu.addSeparator()

            preferencesAction = QAction('Preferences', self)
            preferencesAction.triggered.connect(self.prefs)
            fileMenu.addAction(preferencesAction)
            preferencesAction.setShortcut('Ctrl+P')

            exitAction = QAction('Exit', self)
            exitAction.triggered.connect(self.close)
            exitAction.setShortcut('Ctrl+Q')
            fileMenu.addAction(exitAction)

            extractElementsAction = QAction('Extract Web Elements', self)
            extractElementsAction.triggered.connect(self.extractWebElements)
            extractElementsAction.setShortcut('Ctrl+E')
            toolsMenu.addAction(extractElementsAction)

            sequencerAction = QAction('Create Sequence', self)
            sequencerAction.triggered.connect(self.showSequencer)
            sequencerAction.setShortcut('Ctrl+Shift+S')
            toolsMenu.addAction(sequencerAction)

            scriptEditorAction = QAction('Script Editor', self)
            scriptEditorAction.triggered.connect(self.showScriptEditor)
            toolsMenu.addAction(scriptEditorAction)
            scriptEditorAction.setShortcut('Ctrl+Shift+E')
            scriptEditorAction.setDisabled(True)

            helpAction = QAction('Setup Drivers', self)
            helpAction.triggered.connect(self.showHelpDialog)
            helpMenu.addAction(helpAction)

            howToUseAction = QAction('How to use?', self)
            howToUseAction.triggered.connect(self.howToUseDialog)
            helpMenu.addAction(howToUseAction)

            helpMenu.addSeparator()

            aboutAction = QAction('About', self)
            aboutAction.triggered.connect(self.showAboutDialog)
            helpMenu.addAction(aboutAction)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while setting up menu bar: {e}")

    def addStep(self):
        # STEPS
        try:
            action = self.actionSelection.currentText()
            locator_type = self.locatorSelection.currentText()
            locator_value = self.locatorInput.text()
            text_value = self.inputText.text()
            description_value = self.inputDescription.text()
            sleep_value = self.sleepInput.text()

            if action == 'Sleep':
                step = (action, sleep_value)
                display_txt = f'Sleep for {sleep_value} seconds.'
                self.logger.info(f"Added step: {display_txt}")
            elif action in ['Click Element', 'Input Text']:
                step = (action, locator_type, locator_value, text_value, description_value)
                display_txt = f'{action}: (By: {locator_type if locator_type != "Select Locator" else "N/A"}, {locator_value}){", Text: " + text_value if text_value else ""}, Description: {description_value}'
                self.logger.info(f"Added step: {display_txt}")
            elif action in ['Navigate to URL', 'Execute JavaScript', 'Execute Python Script']:
                step = (action, text_value, description_value)
                display_txt = f'{action}: {text_value}{"." if not description_value else f", Description: {description_value}"}'
                self.logger.info(f"Added step: {display_txt}")
            elif action == 'Maximize Window':
                step = (action,)
                display_txt = f'Maximize Window.'
                self.logger.info(f"Added step: {display_txt}")
            elif action == 'Take Screenshot':
                screenshot_filename = self.inputText.text()
                if not screenshot_filename.endswith('.png'):
                    screenshot_filename += '.png'
                step = (action, screenshot_filename)
                display_txt = f'Take screenshot and save as {screenshot_filename}'
                self.logger.info(f"Added step: {display_txt}")
            elif action == 'Compare Images':
                ref_img_path = self.refImgPath.text()
                test_img_path = self.testImgPath.text()
                output_path = self.outputPath.text()
                step = (action, ref_img_path, test_img_path, output_path)
                display_txt = f'Compare images: Reference: {ref_img_path}, Test: {test_img_path}, Output: {output_path}'
                self.logger.info(f"Added step: {display_txt}")
            else:
                QMessageBox.warning(self, "Invalid Action", "The selected action is not supported.")
                return

            self.steps.append(step)
            self.stepsList.addItem(display_txt)
            self.locatorInput.clear()
            self.inputText.clear()
            self.sleepInput.clear()
            self.inputDescription.clear()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while adding step: {e}")

    def createCheckbox(self, label, tooltip):
        try:
            checkbox = QCheckBox(label, self)
            checkbox.setChecked(False)
            checkbox.setToolTip(tooltip)
            return checkbox
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while creating checkbox: {e}")

    def setupActionSelection(self, layout):
        try:
            self.actionSelection = QComboBox(self)
            actions = ['Select Action', 'Navigate to URL', 'Click Element', 'Input Text', 'Take Screenshot',
                       'Execute JavaScript', 'Sleep', 'Execute Python Script', 'Maximize Window', 'Compare Images']
            self.actionSelection.addItems(actions)
            self.actionSelection.currentIndexChanged.connect(self.updateFields)

            actionSelectionLayout = QVBoxLayout()
            actionSelectionLayout.addWidget(QLabel('Select Action:'))
            actionSelectionLayout.addWidget(self.actionSelection)

            self.locatorSelection = QComboBox(self)
            locator_types = ['Select Locator', 'XPath', 'CSS Selector', 'ID', 'Name', 'Class Name', 'Tag Name',
                             'Link Text',
                             'Partial Link Text']
            self.locatorSelection.addItems(locator_types)

            self.locatorInput = QLineEdit(self)
            self.locatorInput.setPlaceholderText("Enter locator value")

            locatorLayout = QHBoxLayout()
            locatorLayout.addWidget(self.locatorSelection)
            locatorLayout.addWidget(self.locatorInput)

            actionSelectionLayout.addLayout(locatorLayout)

            self.inputText = QLineEdit(self)
            self.inputText.setPlaceholderText("Enter Text")

            self.sleepInput = QLineEdit(self)
            self.sleepInput.setPlaceholderText("Enter Sleep Time (in seconds)")

            self.inputDescription = QLineEdit(self)
            self.inputDescription.setPlaceholderText("Enter Description")

            self.refImgPath = QLineEdit(self)
            self.refImgPath.setToolTip("Enter the path of the reference image to compare")
            self.refImgPath.setPlaceholderText("Enter Reference Image Path")

            self.testImgPath = QLineEdit(self)
            self.testImgPath.setToolTip("Enter the name of the image, with '.png', to compare with the reference image")
            self.testImgPath.setPlaceholderText("Enter File Name")

            self.outputPath = QLineEdit(self)
            self.outputPath.setToolTip("Enter the path where the output image should be saved")
            self.outputPath.setPlaceholderText("Enter Output Path")

            self.openPhoto = QCheckBox("Open Photo When Done", self)
            self.openPhoto.setToolTip("Open the photo after comparing images")
            self.openPhoto.setChecked(False)

            fieldsLayout = QVBoxLayout()
            fieldsLayout.addWidget(self.inputText)
            fieldsLayout.addWidget(self.sleepInput)
            fieldsLayout.addWidget(self.inputDescription)
            fieldsLayout.addWidget(self.refImgPath)
            fieldsLayout.addWidget(self.testImgPath)
            fieldsLayout.addWidget(self.outputPath)
            fieldsLayout.addWidget(self.openPhoto)

            self.refImgPath.setVisible(False)
            self.testImgPath.setVisible(False)
            self.outputPath.setVisible(False)
            self.openPhoto.setVisible(False)

            actionSelectionLayout.addLayout(fieldsLayout)
            layout.addLayout(actionSelectionLayout)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while setting up action selection: {e}")

    def setupButtonsAndStepsList(self, layout):
        try:
            self.editMode = False
            self.editIndex = None

            self.generateReport = QCheckBox("Generate Report", self)
            self.generateReport.setChecked(False)
            self.generateReport.setToolTip("Generate a report after the automation is completed")
            self.generateReport.setStyleSheet("padding-top: 10px;padding-bottom: 10px;")
            layout.addWidget(self.generateReport)

            self.addButton = QPushButton('Add Step', self)
            self.addButton.clicked.connect(self.addOrEditStep)

            self.removeButton = QPushButton('Remove Selected Step', self)
            self.removeButton.clicked.connect(self.removeSelectedStep)

            self.editButton = QPushButton('Edit Selected Step', self)
            self.editButton.clicked.connect(self.editSelectedStep)

            self.saveButton = QPushButton('Save Changes', self)
            self.saveButton.clicked.connect(self.saveEditedStep)

            self.saveButton.setEnabled(True)
            self.saveButton.setVisible(False)

            self.stepsList = QListWidget(self)

            self.startButton = QPushButton('Run', self)
            self.startButton.clicked.connect(self.startAutomation)

            # self.stopButton = QPushButton('Stop', self)
            # self.stopButton.clicked.connect(self.stopAutomation)
            # self.stopButton.setEnabled(False)
            # self.stopButton.setVisible(False)
            # self.stopButton.setProperty("class", "red-btn")

            self.moveUpButton = QPushButton('Up', self)
            self.moveDownButton = QPushButton('Down', self)
            self.moveUpButton.clicked.connect(self.moveStepUp)
            self.moveDownButton.clicked.connect(self.moveStepDown)

            self.moveUpButton.setProperty("class", "secondary-btn")
            self.moveDownButton.setProperty("class", "secondary-btn")

            self.locatorSelection.setVisible(False)
            self.locatorInput.setVisible(False)
            self.inputText.setVisible(False)
            self.sleepInput.setVisible(False)
            self.inputDescription.setVisible(False)

            buttonsLayout = QHBoxLayout()
            buttonsLayout.addWidget(self.addButton)
            buttonsLayout.addWidget(self.editButton)
            buttonsLayout.addWidget(self.saveButton)
            buttonsLayout.addWidget(self.removeButton)
            buttonsLayout.addWidget(self.moveUpButton)
            buttonsLayout.addWidget(self.moveDownButton)
            buttonsLayout.addWidget(self.startButton)
            # buttonsLayout.addWidget(self.stopButton)

            self.startButton.setStyleSheet("""
                QPushButton {
                    color: white;
                    background-color: #28A745;
                    border-radius: 4px;
                    padding: 6px 12px;
                    border: none;
                    font-size: 12px;
                }

                QPushButton:hover {
                    background-color: #218838;
                    border-color: #1E7E34;
                }

                QPushButton:pressed {
                    background-color: #1D7D33;
                    border-color: #1C7430;
                }
            """)
            # self.stopButton.setStyleSheet("""
            #     QPushButton {
            #         color: white;
            #         background-color: #DC3545;
            #         border-radius: 4px;
            #         padding: 6px 12px;
            #         border: none;
            #         font-size: 12px;
            #     }
            #
            #     QPushButton:hover {
            #         background-color: #C82333;
            #         border-color: #BD2130;
            #     }
            #
            #     QPushButton:pressed {
            #         background-color: #B21F2D;
            #         border-color: #B21F2D;
            #     }
            # """)

            layout.addLayout(buttonsLayout)
            layout.addWidget(self.stepsList)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while setting up buttons and steps list: {e}")

    def updateLocatorFields(self):
        try:
            locator_type = self.locatorSelection.currentText()
            self.locatorInput.setVisible(locator_type != 'Select Locator')
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating locator fields: {e}")

    def removeSelectedStep(self):
        try:
            selected_item = self.stepsList.currentRow()
            if selected_item >= 0:
                del self.steps[selected_item]
                self.stepsList.takeItem(selected_item)
                self.logger.info(f"Removed step at index {selected_item}.")
            else:
                QMessageBox.warning(self, "No Selection", "Please select a step to remove.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while removing selected step: {e}")

    def updateFields(self):
        try:
            action = self.actionSelection.currentText()

            if action in ['Click Element', 'Input Text']:
                self.locatorSelection.setVisible(True)
                self.locatorInput.setVisible(True)
            else:
                self.locatorSelection.setVisible(False)
                self.locatorInput.setVisible(False)

            self.inputText.setVisible(
                action in ['Input Text', 'Execute Python Script', 'Execute JavaScript', 'Navigate to URL',
                           'Take Screenshot'])
            self.sleepInput.setVisible(action == 'Sleep')
            self.inputDescription.setVisible(
                action in ['Click Element', 'Input Text', 'Sleep', 'Navigate to URL', 'Execute Python Script',
                           'Execute JavaScript'])
            self.refImgPath.setVisible(action == 'Compare Images')
            self.testImgPath.setVisible(action == 'Compare Images')
            self.outputPath.setVisible(action == 'Compare Images')
            self.openPhoto.setVisible(action == 'Compare Images')

            if action == 'Navigate to URL':
                self.inputText.setPlaceholderText("Enter URL")
            elif action == 'Input Text':
                self.inputText.setPlaceholderText("Enter Text")
            elif action == 'Execute Python Script':
                self.inputText.setPlaceholderText("Enter Script Path")
            elif action == 'Execute JavaScript':
                self.inputText.setPlaceholderText("Enter JavaScript Code")
            elif action == 'Sleep':
                self.inputText.setPlaceholderText("Enter Sleep Time (in seconds)")
            elif action == 'Take Screenshot':
                self.inputText.setPlaceholderText("Enter Screenshot Name")
            else:
                self.inputText.setPlaceholderText("Enter Text")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating fields: {e}")

    class ResultsWindow(QDialog):
        def __init__(self, parent=None):
            try:
                super().__init__(parent)
                self.setWindowTitle(
                    f"Results for {parent.testName.text() if parent.testName.text() else 'Unnamed Test'}")
                self.setGeometry(100, 100, 600, 600)

                self.testNameLabel = QLabel(parent.testName.text() if parent.testName.text() else "Unnamed Test", self)
                self.testNameLabel.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 10px;")
                self.testNameLabel.setMaximumWidth(550)

                self.testDescriptionLabel = QLabel(
                    parent.testDescription.text() if parent.testDescription.text() else "No description provided.",
                    self)
                self.testDescriptionLabel.setStyleSheet("font-size: 12px; margin-bottom: 10px;")
                self.testDescriptionLabel.setWordWrap(True)
                self.testDescriptionLabel.setMaximumWidth(550)

                if any(checkbox.isChecked() for checkbox in parent.findChildren(QCheckBox)):
                    self.browserOptionsLabel = QLabel("Performed on Chrome, with the following options:", self)
                    self.browserOptionsLabel.setWordWrap(True)
                    self.browserOptionsLabel.setMaximumWidth(550)
                    for checkbox in parent.findChildren(QCheckBox):
                        if checkbox.isChecked():
                            if checkbox.text() == "Generate Report":
                                continue
                            self.browserOptionsLabel.setText(f"{self.browserOptionsLabel.text()}\n - {checkbox.text()}")
                        else:
                            continue
                else:
                    self.browserOptionsLabel = QLabel("Performed on Chrome, with no options selected.", self)
                    self.browserOptionsLabel.setWordWrap(True)
                    self.browserOptionsLabel.setMaximumWidth(550)

                self.resultsTable = QTableWidget(self)
                self.resultsTable.setColumnCount(2)
                self.resultsTable.setHorizontalHeaderLabels(["Step", "Status"])
                self.resultsTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
                self.resultsTable.verticalHeader().setVisible(False)
                self.resultsTable.setEditTriggers(QTableWidget.NoEditTriggers)
                self.resultsTable.setAlternatingRowColors(True)
                self.resultsTable.setShowGrid(False)

                self.resultsTable.setStyleSheet("""
                    QTableWidget {
                        border: 1px solid #ddd;
                        border-radius: 4px;
                        color: #555;
                        background-color: #f5f5f5;
                    }
                    QTableWidget::item {
                        padding: 4px;
                        color: #555;
                    }
                    QTableWidget::item:selected {
                        background-color: #007BFF;
                        color: white;
                    }
                """)

                self.exportButton = QPushButton('Export to Excel', self)
                self.exportButton.clicked.connect(self.exportReport)

                self.buttonGroupBox = QGroupBox("Export Descriptions to", self)
                buttonGroupLayout = QHBoxLayout()

                self.copyJiraMarkdownButton = QPushButton('JIRA', self.buttonGroupBox)
                self.copyJiraMarkdownButton.clicked.connect(self.copyJiraMarkdown)
                self.copyJiraMarkdownButton.setProperty("class", "secondary-btn")
                buttonGroupLayout.addWidget(self.copyJiraMarkdownButton)

                self.mondayButton = QPushButton('Monday.com', self.buttonGroupBox)
                self.mondayButton.setProperty("class", "secondary-btn")
                buttonGroupLayout.addWidget(self.mondayButton)

                self.proofhubButton = QPushButton('ProofHub', self.buttonGroupBox)
                self.proofhubButton.setProperty("class", "secondary-btn")
                buttonGroupLayout.addWidget(self.proofhubButton)

                self.clickupButton = QPushButton('ClickUp', self.buttonGroupBox)
                self.clickupButton.setProperty("class", "secondary-btn")
                buttonGroupLayout.addWidget(self.clickupButton)

                self.buttonGroupBox.setLayout(buttonGroupLayout)

                layout = QVBoxLayout(self)
                layout.addWidget(self.testNameLabel)
                layout.addWidget(self.testDescriptionLabel)
                layout.addWidget(self.browserOptionsLabel)
                layout.addWidget(self.resultsTable)
                layout.addWidget(self.buttonGroupBox)
                layout.addWidget(self.exportButton)
                self.setLayout(layout)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Error while initializing results window: {e}")

        def copyJiraMarkdown(self):
            try:
                jiraMarkdown = self.parent().generateBugForJira()
                clipboard = QApplication.clipboard()
                clipboard.setText(jiraMarkdown)
                QMessageBox.information(self, "Copied", "JIRA Markdown copied to clipboard.")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Error while copying JIRA Markdown: {e}")

        def exportReport(self):
            try:
                data = []

                # Adding test details
                data.append(["Test Name", self.parent().testName.text()])
                data.append(["Description", self.parent().testDescription.text()])
                data.append(["", ""])  # Spacer row

                # Collecting steps and statuses
                for row in range(self.resultsTable.rowCount()):
                    data.append([self.resultsTable.item(row, 0).text(), self.resultsTable.item(row, 1).text()])
                data.append(["", ""])  # Spacer row

                df = pd.DataFrame(data, columns=["Step", "Status"])

                # Collecting browser options
                browser_options = [checkbox.text() for checkbox in self.parent().findChildren(QCheckBox) if
                                   checkbox.isChecked() and checkbox.text() != "Generate Report"]
                browser_options_row = ["Browser Options", ', '.join(browser_options)] if browser_options else [
                    "Browser Options", 'None']
                df.loc[df.index.max() + 1] = browser_options_row

                # Create a new Excel workbook
                wb = Workbook()
                ws = wb.active

                # Convert DataFrame to Excel rows
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

                        # Applying styles
                        if r_idx == 0:  # Header row
                            cell.font = Font(bold=True)
                        if c_idx == 2:  # Status column
                            if value == "Passed":
                                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                            elif value == "Failed":
                                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

                # Save the workbook
                wb.save(f"{self.parent().testName.text()}.xlsx")
                QMessageBox.information(self, "Exported", "Report exported successfully.")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Error while exporting report: {e}")

    def startAutomation(self):
        # self.startButton.setEnabled(False)
        # self.startButton.setVisible(False)
        # self.stopButton.setEnabled(True)
        # self.stopButton.setVisible(True)
        self.logger.info("Starting automation...")
        try:
            chromeOptionsMapping = {
                "Headless Mode": "--headless",
                "Disable GPU": "--disable-gpu",
                "Incognito Mode": "--incognito",
                "Disable Popup Blocking": "--disable-popup-blocking",
                "Disable Infobars": "--disable-infobars",
                "Disable Extensions": "--disable-extensions",
                "Disable Dev Shm Usage": "--disable-dev-shm-usage",
                "Ignore Certificate Errors": "--ignore-certificate-errors",
                "Custom User Agent": "--user-agent",
                "Disable JavaScript": "--disable-javascript",
                "Disable Images": "--blink-settings=imagesEnabled=false",
                "Enable Network Throttling": "--enable-network-throttling",
                "Enable Performance Logging": "--enable-performance-logging",
                "Enable GPU Hardware Acceleration": "--enable-gpu-rasterization",
                "Remote Debugging Port": "--remote-debugging-port",
                "Proxy Settings": "--proxy-server",
                "Enable Automation": "--enable-automation",
                "No Sandbox": "--no-sandbox",
                "Disable Web Security": "--disable-web-security",
                "Enable Experimental Features": "--enable-experimental-web-platform-features",
                "Disable Password Manager": "--disable-password-manager-reauthentication",
                "Disable Autofill": "--disable-autofill-keyboard-accessory-view",
                "Disable Filesystem API": "--disable-filesystem",
                "Disable Geolocation": "--disable-geolocation",
            }

            edgeOptionsMapping = {
                "Headless Mode": "headless",
                "Disable GPU": "disable-gpu",
                "InPrivate Mode": "InPrivate",
                "Disable Popup Blocking": "disable-popup-blocking",
                "Disable Extensions": "disable-extensions",
                "Ignore Certificate Errors": "ignore-certificate-errors",
                "Custom User Agent": "user-agent",
                "Disable JavaScript": "disable-javascript",
                "Disable Images": "disable-images",
                "Enable Network Throttling": "enable-network-throttling",
                "Enable Performance Logging": "enable-performance-logging",
                "Enable GPU Hardware Acceleration": "enable-gpu-rasterization",
                "Remote Debugging Port": "remote-debugging-port",
                "Proxy Settings": "proxy-server",
                "Enable Automation": "enable-automation",
                "No Sandbox": "no-sandbox",
                "Disable Web Security": "disable-web-security",
                "Enable Experimental Features": "enable-experimental-web-platform-features",
                "Disable Password Manager": "disable-password-manager",
                "Disable Autofill": "disable-autofill",
                "Disable Filesystem API": "disable-filesystem",
                "Disable Geolocation": "disable-geolocation",
            }

            browser_type = self.loadSetting("defaultBrowser", "Chrome")

            chrome_driver_location = self.loadSetting("driverLocation", "chromedriver.exe")
            msedge_driver_location = self.loadSetting("msedgeLocation", "msedgedriver.exe")

            try:
                if browser_type == "Chrome":
                    chrome_options = Options()
                    for checkbox in self.findChildren(QCheckBox):
                        selenium_option = chromeOptionsMapping.get(checkbox.text())
                        if selenium_option and checkbox.isChecked():
                            chrome_options.add_argument(selenium_option)

                    if not os.path.isfile(chrome_driver_location):
                        raise ValueError("Invalid Chrome driver location")
                    self.logger.info("Starting Chrome browser with WebDriver at: " + chrome_driver_location)
                    self.driver = webdriver.Chrome(chrome_options)
                elif browser_type == "Edge":
                    edge_options = EdgeOptions()
                    for checkbox in self.findChildren(QCheckBox):
                        selenium_option = edgeOptionsMapping.get(checkbox.text())
                        if selenium_option and checkbox.isChecked():
                            edge_options.add_argument(selenium_option)

                    if not os.path.isfile(msedge_driver_location):
                        raise ValueError("Invalid Edge driver location")
                    self.logger.info("Starting Edge browser with WebDriver at: " + msedge_driver_location)
                    self.driver = webdriver.Edge(edge_options)
                # Add support for other browsers here
                else:
                    raise ValueError("Unsupported browser type")

                locator_strategies = {
                    'XPath': By.XPATH,
                    'CSS Selector': By.CSS_SELECTOR,
                    'ID': By.ID,
                    'Name': By.NAME,
                    'Class Name': By.CLASS_NAME,
                    'Tag Name': By.TAG_NAME,
                    'Link Text': By.LINK_TEXT,
                    'Partial Link Text': By.PARTIAL_LINK_TEXT
                }

                if self.generateReport.isChecked():
                    results_window = self.ResultsWindow(self)
                    results_window.resultsTable.setRowCount(len(self.steps))

                self.results = []

                for step in self.steps:
                    action = step[0]
                    try:
                        # STEP
                        if action == 'Navigate to URL':
                            try:
                                self.driver.get(step[1])
                            except Exception as e:
                                self.logger.error(f"Error while navigating to URL: {e}")
                        elif action in ['Click Element', 'Input Text']:
                            try:
                                locator_type = step[1]
                                locator_value = step[2]
                                element = self.driver.find_element(locator_strategies[locator_type], locator_value)
                                if action == 'Click Element':
                                    element.click()
                                else:
                                    element.send_keys(step[3])
                                self.logger.info(f"{action} at {locator_type}: {locator_value}")
                            except Exception as e:
                                self.logger.error(f"Error while performing {action}: {e}")
                        elif action == 'Take Screenshot':
                            try:
                                with open(self.settingsFilePath(), "r") as settings_file:
                                    settings = json.load(settings_file)
                                    screenshot_folder = settings.get("savePath")
                                    if not os.path.isdir(screenshot_folder):
                                        os.makedirs(screenshot_folder)
                                    screenshot_filename = os.path.join(screenshot_folder, step[1])
                                    self.driver.save_screenshot(screenshot_filename)
                                    self.logger.info(f"Screenshot saved as {screenshot_filename}")
                            except Exception as e:
                                self.logger.error(f"Error while taking screenshot: {e}")
                        elif action == 'Execute JavaScript':
                            try:
                                self.driver.execute_script(step[1])
                                self.logger.info(f"Executed JavaScript: {step[1]}")
                            except Exception as e:
                                self.logger.error(f"Error in JavaScript: {e}")
                        elif action == 'Sleep':
                            try:
                                time.sleep(float(step[1]))
                                self.logger.info(f"Slept for {step[1]} seconds.")
                            except Exception as e:
                                self.logger.error(f"Error while sleeping: {e}")
                        elif action == 'Maximize Window':
                            try:
                                self.driver.maximize_window()
                                self.logger.info("Maximized window.")
                            except Exception as e:
                                self.logger.error(f"Error while maximizing window: {e}")
                        elif action == 'Execute Python Script':
                            try:
                                exec(open(step[1]).read())
                                self.logger.info(f"Executed Python script: {step[1]}")
                            except Exception as e:
                                self.logger.error(f"Error in Python script: {e}")
                        elif action == 'Compare Images':
                            try:
                                self.logger.info(f"Comparing images: {step[1]} and {step[2]}")
                                if self.compareImages(step[1], step[2], step[3]):
                                    self.results.append((step, 'Passed'))
                                    if self.openPhoto.isChecked():
                                        os.startfile(step[3])
                                else:
                                    self.results.append((step, 'Failed'))
                            except Exception as e:
                                self.logger.error(f"Error in {action}: {e}")
                                self.results.append((step, 'Failed'))

                        self.results.append((step, 'Passed'))

                    except Exception as e:
                        self.logger.error(f"Error in {action}: {e}")
                        self.results.append((step, 'Failed'))

                self.driver.quit()

                if self.generateReport.isChecked():
                    self.displayResults(self.results)

                self.logger.info("\n\nOperation completed successfully.\n")

            except Exception as e:
                self.logger.error(f"Error: {e}")
                QMessageBox.critical(self, "Error", str(e))
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while running the script: {e}")

    # def stopAutomation(self):
    #     self.startButton.setEnabled(True)
    #     self.startButton.setVisible(True)
    #     self.stopButton.setEnabled(False)
    #     self.stopButton.setVisible(False)
    #     try:
    #         self.driver.quit()
    #         self.logger.info("Operation stopped by user.")
    #         self.logger.info("\n\nOperation stopped by user.\n")
    #     except Exception as e:
    #         QMessageBox.warning(self, "Error", f"Error while stopping the script: {e}")

    def formatStepText(self, step):
        # STEP
        try:
            action = step[0]
            if action == 'Sleep':
                return f'Sleep for {step[1]} seconds'
            elif action in ['Click Element', 'Input Text']:
                return f'{action} at {step[1]}: {step[2]}'
            elif action in ['Navigate to URL', 'Execute JavaScript', 'Execute Python Script']:
                return f'{action}: {step[1]}'
            elif action == 'Take Screenshot':
                return f'Take screenshot: {step[1]}'
            elif action == 'Maximize Window':
                return 'Maximize Window'
            elif action == 'Compare Images':
                return f'Comparing images: {step[1]} and {step[2]}'
            else:
                return 'Unknown Action'
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while formatting step text: {e}")

    def displayResults(self, results):
        try:
            results_window = self.ResultsWindow(self)
            results_window.resultsTable.setRowCount(len(results))

            for index, (step, status) in enumerate(results):
                step_text = self.formatStepText(step)
                step_item = QTableWidgetItem(step_text)
                status_item = QTableWidgetItem(status)

                if status == 'Passed':
                    status_item.setBackground(QColor(203, 255, 171))
                else:
                    status_item.setBackground(QColor(255, 171, 171))

                results_window.resultsTable.setItem(index, 0, step_item)
                results_window.resultsTable.setItem(index, 1, status_item)

            results_window.exec_()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while displaying results: {e}")

    def setupLogging(self):
        try:
            self.logger = logging.getLogger('Atom8')
            logging.basicConfig(level=logging.INFO)

            logTextBox = QTextEditLogger(self.logViewer)
            logTextBox.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
            logging.getLogger().addHandler(logTextBox)
            logging.getLogger().setLevel(logging.DEBUG)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while setting up logging: {e}")

    def showAboutDialog(self):
        try:
            QMessageBox.about(self, "About Atom8", """
        <html>
        <head>
            <style> 
                p { font-family: Arial, sans-serif; line-height: 1.0; }
                a { text-decoration: none; color: #007BFF; }
                a:hover { text-decoration: underline; }
            </style>
        </head>
        <body>
            <h2>Atom8</h2>
            <p><strong>Version:</strong> %s-dev b%s</p>
            <p>Atom8 is a robust and user-friendly web automation platform, offering enhanced capabilities for both professionals and enthusiasts. This tool streamlines complex web tasks, providing an advanced yet seamless automation experience. It's perfect for a variety of applications, including data scraping, automated testing, and more.</p>
            <p>Built upon the popular Selenium framework for basic operations, Atom8 stands out as a more accessible alternative, boasting a straightforward interface for creating and executing both simple and complex automation scripts.</p>
            <p>Explore more about Atom8, get the latest updates, and access support on our GitHub page: <a href="https://github.com/Dcohen52/Atom8" target="_blank">Atom8 GitHub Repository</a>.</p>
            <p><strong>Created by:</strong> Dekel Cohen</p>
            <p><strong>License:</strong> MIT License</p>
            <p><strong>Disclaimer:</strong> Atom8 is an independent project and is not officially affiliated with or endorsed by the Selenium project or its associates.</p>
        </body>
        </html>
        """ % (__version__, __build__))
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing about dialog: {e}")

    def showHelpDialog(self):
        try:
            QMessageBox.about(self, "Setup Drivers", """
        <html>
        <head>
            <style>
                p { font-family: Arial, sans-serif; line-height: 1.0; }
                a { text-decoration: none; color: #007BFF; }
                a:hover { text-decoration: underline; }
            </style>
        </head>
        <body>
            <h2>How to setup drivers?</h2>
            <p>Drivers are required to run the browser. You can download the drivers from the following links:</p>
            <p><strong>Chrome:</strong> <a href="https://chromedriver.chromium.org/downloads" target="_blank">Chrome Driver</a></p>
            <p><strong>Edge:</strong> <a href="https://developer.microsoft.com/en-us/microsoft-edge/tools/webdriver/" target="_blank">Edge Driver</a></p>
            <p><strong>Firefox:</strong> <a href="https://github.com/mozilla/geckodriver/releases" target="_blank">Firefox Driver</a></p>
            <p><strong>Safari:</strong> <a href="https://developer.apple.com/documentation/webkit/testing_with_webdriver_in_safari" target="_blank">Safari Driver Help</a></p>
            <p><strong>Drivers Location:</strong> After downloading the driver, you need to specify the location of the driver in the preferences menu.</p>
        </body>
        </html>
        """)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing help dialog: {e}")

    def howToUseDialog(self):
        try:
            QMessageBox.about(self, "How to use", """
        <html>
        <head>
            <style>
                p { font-family: Arial, sans-serif; line-height: 1.0; }
                a { text-decoration: none; color: #007BFF; }
                a:hover { text-decoration: underline; }
            </style>
        </head>
        <body>
            <h2>How to use</h2>
            <p><strong>Build Script:</strong> You can build the script by adding steps to the steps list.</p>
            <p><strong>Run Script:</strong> Click the Run button to start the automation process.</p>
            <p><strong>Sequencer:</strong> You can execute multiple atm8 files in sequence by adding them to the sequencer.</p>
            <h3>Steps</h3>
            <p><strong>Navigate to URL:</strong> Navigate to a specific URL.</p>
            <p><strong>Click Element:</strong> Click on an element on the page.</p>
            <p><strong>Input Text:</strong> Input text into an element on the page.</p>
            <p><strong>Take Screenshot:</strong> Take a screenshot of the current page.</p>
            <p><strong>Execute JavaScript:</strong> Execute JavaScript code on the page.</p>
            <p><strong>Execute Python Script:</strong> Execute a Python script.</p>
            <p><strong>Sleep:</strong> Pause the script for a specified amount of time.</p>
            <p><strong>Maximize Window:</strong> Maximize the browser window.</p>
            <p><strong>Compare Images:</strong> Compare two images.</p>
            <hr>
            <p>For full documentation, information about available options and usage - visit the <a href="https://github.com/Dcohen52/Atom8" target="_blank">Atom8 GitHub Repository</a>.</p>
        </body>
        </html>
        """)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing how to use dialog: {e}")

    def saveFile(self):
        try:
            fileName, _ = QFileDialog.getSaveFileName(self, "Save As", "", "Atom8 Files (*.atm8)")
            if fileName:
                self.currentFilePath = fileName
                self.updateRecentFiles(fileName)
                with open(fileName, "w+") as file:
                    file_content = {
                        "testName": self.testName.text(),
                        "testDescription": self.testDescription.text(),
                        "steps": self.steps
                    }
                    json.dump(file_content, file)
                self.statusBar.showMessage(f"File saved as {fileName} successfully.", 5000)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving file: {e}")

    def realSaveFile(self):
        try:
            if self.currentFilePath:
                with open(self.currentFilePath, "w") as file:
                    file_content = {
                        "testName": self.testName.text(),
                        "testDescription": self.testDescription.text(),
                        "steps": self.steps
                    }
                    json.dump(file_content, file)
                self.statusBar.showMessage(f"File {self.currentFilePath} saved successfully.", 5000)
            else:
                self.saveFile()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving file: {e}")

    def openFile(self):
        try:
            fileName, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Atom8 Files (*.atm8)")
            if fileName:
                self.currentFilePath = fileName
                self.updateRecentFiles(fileName)
                try:
                    with open(fileName, "r") as file:
                        file_content = json.load(file)
                        if not isinstance(file_content, dict) or "steps" not in file_content:
                            raise ValueError("File content is not in the expected format")

                        self.testName.setText(file_content.get("testName", ""))
                        self.testDescription.setText(file_content.get("testDescription", ""))
                        self.steps = file_content["steps"]
                        self.stepsList.clear()
                        for index, step in enumerate(self.steps):
                            if not isinstance(step, (list, tuple)):
                                QMessageBox.critical(self, "Error", f"Invalid step format at index {index}: {step}")
                                continue
                            display_text = self.constructStepDisplayText(step)
                            self.stepsList.addItem(display_text)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to open file: {e}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while opening file: {e}")

    def constructStepDisplayText(self, step):
        # STEP
        try:
            action = step[0]
            if action == 'Sleep':
                display_text = f'Sleep for {step[1]} seconds.'
            elif action in ['Click Element', 'Input Text']:
                display_text = f'{action}: (By: {step[1]}, {step[2]}){", Text: " + step[3] if step[3] else ""}, Description: {step[4]}'
            elif action in ['Navigate to URL', 'Execute JavaScript', 'Execute Python Script']:
                display_text = f'{action}: {step[1]}{"." if not step[2] else f", Description: {step[2]}"}'
            elif action == 'Take Screenshot':
                display_text = f'Take screenshot and save as {step[1]}'
            elif action == 'Compare Images':
                display_text = f'Comparing images: {step[1]} and {step[2]}'
            else:
                display_text = f'{action}'
        except Exception as e:
            display_text = f'Error: {e}'
        return display_text

    def clearStepsList(self):
        self.stepsList.clear()
        self.steps.clear()
        self.clearInputFields()
        self.clearLogs()

    def clearAllFields(self):
        self.testName.clear()
        self.testDescription.clear()
        self.clearStepsList()
        self.clearInputFields()
        self.clearLogs()
        # clear the saved file path
        self.currentFilePath = None

    def prefs(self):
        try:
            self.prefsWindow = QDialog(self, Qt.Window)
            self.prefsWindow.setWindowTitle("Preferences")
            prefsLayout = QVBoxLayout()

            # Create a tab widget for different preference sections
            tabWidget = QTabWidget()
            prefsLayout.addWidget(tabWidget)

            # General settings tab
            generalTab = QWidget()
            generalLayout = QVBoxLayout()

            browserLabel = QLabel("Default Browser:")
            self.browserComboBox = QComboBox()
            self.browserComboBox.addItems(["Chrome", "Firefox", "Safari", "Edge"])
            self.browserComboBox.setCurrentText(self.loadSetting("defaultBrowser", "Chrome"))
            browserLayout = QHBoxLayout()
            browserLayout.addWidget(browserLabel)
            browserLayout.addWidget(self.browserComboBox)
            generalLayout.addLayout(browserLayout)

            savePathLabel = QLabel("Default Screenshots Save Path:")
            self.savePathLineEdit = QLineEdit()
            current_save_path = self.loadSetting("savePath", "")
            if current_save_path:
                self.savePathLineEdit.setText(current_save_path)
            savePathButton = QPushButton("Choose")
            savePathButton.clicked.connect(self.chooseSavePathLocation)
            savePathLayout = QHBoxLayout()
            savePathLayout.addWidget(savePathLabel)
            savePathLayout.addWidget(self.savePathLineEdit)
            savePathLayout.addWidget(savePathButton)
            generalLayout.addLayout(savePathLayout)

            driverLocationLabel = QLabel("Chrome Driver Location:")
            self.driverLocationLineEdit = QLineEdit()
            self.driverLocationLineEdit.setText(self.loadSetting("driverLocation", ""))
            driverLocationButton = QPushButton("Choose")
            driverLocationButton.clicked.connect(self.chooseChromeDriverLocation)
            driverLocationLayout = QHBoxLayout()
            driverLocationLayout.addWidget(driverLocationLabel)
            driverLocationLayout.addWidget(self.driverLocationLineEdit)
            driverLocationLayout.addWidget(driverLocationButton)
            generalLayout.addLayout(driverLocationLayout)

            msedgeLocationLabel = QLabel("Edge Driver Location:")
            self.msedgeLocationLineEdit = QLineEdit()
            self.msedgeLocationLineEdit.setText(self.loadSetting("msedgeLocation", ""))
            msedgeLocationButton = QPushButton("Choose")
            msedgeLocationButton.clicked.connect(self.chooseMsEdgeDriverLocation)
            msedgeLocationLayout = QHBoxLayout()
            msedgeLocationLayout.addWidget(msedgeLocationLabel)
            msedgeLocationLayout.addWidget(self.msedgeLocationLineEdit)
            msedgeLocationLayout.addWidget(msedgeLocationButton)
            generalLayout.addLayout(msedgeLocationLayout)

            generalTab.setLayout(generalLayout)
            tabWidget.addTab(generalTab, "General")

            # ProofHub settings tab
            proofhubTab = QWidget()
            proofhubLayout = QVBoxLayout()

            proofhubAPI = QLabel("ProofHub API Key:")
            self.proofhubAPIKey = QLineEdit()
            self.proofhubAPIKey.setText(self.loadSetting("proofhubAPIKey", ""))
            proofhubAPILayout = QHBoxLayout()
            proofhubAPILayout.addWidget(proofhubAPI)
            proofhubAPILayout.addWidget(self.proofhubAPIKey)
            proofhubLayout.addLayout(proofhubAPILayout)

            proofhubProject = QLabel("ProofHub Project ID:")
            self.proofhubProjectID = QLineEdit()
            self.proofhubProjectID.setText(self.loadSetting("proofhubProjectID", ""))
            proofhubProjectLayout = QHBoxLayout()
            proofhubProjectLayout.addWidget(proofhubProject)
            proofhubProjectLayout.addWidget(self.proofhubProjectID)
            proofhubLayout.addLayout(proofhubProjectLayout)

            proofhubTaskList = QLabel("ProofHub Task List ID:")
            self.proofhubTaskListID = QLineEdit()
            self.proofhubTaskListID.setText(self.loadSetting("proofhubTaskListID", ""))
            proofhubTaskListLayout = QHBoxLayout()
            proofhubTaskListLayout.addWidget(proofhubTaskList)
            proofhubTaskListLayout.addWidget(self.proofhubTaskListID)
            proofhubLayout.addLayout(proofhubTaskListLayout)

            proofhubTab.setLayout(proofhubLayout)
            tabWidget.addTab(proofhubTab, "ProofHub")

            # JIRA settings tab
            jiraTab = QWidget()
            jiraLayout = QVBoxLayout()

            jiraURLLabel = QLabel("JIRA URL:")
            self.jiraURLLineEdit = QLineEdit()
            self.jiraURLLineEdit.setText(self.loadSetting("jiraURL", ""))
            jiraURLLayout = QHBoxLayout()
            jiraURLLayout.addWidget(jiraURLLabel)
            jiraURLLayout.addWidget(self.jiraURLLineEdit)
            jiraLayout.addLayout(jiraURLLayout)

            jiraUserLabel = QLabel("JIRA Username:")
            self.jiraUserLineEdit = QLineEdit()
            self.jiraUserLineEdit.setText(self.loadSetting("jiraUsername", ""))
            jiraUserLayout = QHBoxLayout()
            jiraUserLayout.addWidget(jiraUserLabel)
            jiraUserLayout.addWidget(self.jiraUserLineEdit)
            jiraLayout.addLayout(jiraUserLayout)

            jiraTokenLabel = QLabel("JIRA API Token:")
            self.jiraTokenLineEdit = QLineEdit()
            self.jiraTokenLineEdit.setText(self.loadSetting("jiraAPIToken", ""))
            jiraTokenLayout = QHBoxLayout()
            jiraTokenLayout.addWidget(jiraTokenLabel)
            jiraTokenLayout.addWidget(self.jiraTokenLineEdit)
            jiraLayout.addLayout(jiraTokenLayout)

            jiraTab.setLayout(jiraLayout)
            tabWidget.addTab(jiraTab, "JIRA")

            # Monday settings tab
            mondayTab = QWidget()
            mondayLayout = QVBoxLayout()

            mondayURLLabel = QLabel("Monday URL:")
            self.mondayURLLineEdit = QLineEdit()
            self.mondayURLLineEdit.setText(self.loadSetting("mondayURL", ""))
            mondayURLLayout = QHBoxLayout()
            mondayURLLayout.addWidget(mondayURLLabel)
            mondayURLLayout.addWidget(self.mondayURLLineEdit)
            mondayLayout.addLayout(mondayURLLayout)

            mondayTokenLabel = QLabel("Monday API Token:")
            self.mondayTokenLineEdit = QLineEdit()
            self.mondayTokenLineEdit.setText(self.loadSetting("mondayAPIToken", ""))
            mondayTokenLayout = QHBoxLayout()
            mondayTokenLayout.addWidget(mondayTokenLabel)
            mondayTokenLayout.addWidget(self.mondayTokenLineEdit)
            mondayLayout.addLayout(mondayTokenLayout)

            mondayTab.setLayout(mondayLayout)
            tabWidget.addTab(mondayTab, "Monday")

            # ClickUp settings tab
            clickupTab = QWidget()
            clickupLayout = QVBoxLayout()

            clickupTokenLabel = QLabel("ClickUp API Token:")
            self.clickupTokenLineEdit = QLineEdit()
            self.clickupTokenLineEdit.setText(self.loadSetting("clickupAPIToken", ""))
            clickupTokenLayout = QHBoxLayout()
            clickupTokenLayout.addWidget(clickupTokenLabel)
            clickupTokenLayout.addWidget(self.clickupTokenLineEdit)
            clickupLayout.addLayout(clickupTokenLayout)

            clickupTab.setLayout(clickupLayout)
            tabWidget.addTab(clickupTab, "ClickUp")

            saveButton = QPushButton("Save")
            saveButton.clicked.connect(self.savePrefs)
            prefsLayout.addWidget(saveButton)

            self.prefsWindow.setLayout(prefsLayout)
            self.prefsWindow.resize(600, 300)
            self.prefsWindow.show()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while opening preferences: {e}")

    def chooseChromeDriverLocation(self):
        try:
            fileName, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Executable Files (*.exe)")
            if fileName:
                self.driverLocationLineEdit.setText(fileName)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while choosing Chrome driver location: {e}")

    def chooseMsEdgeDriverLocation(self):
        try:
            fileName, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Executable Files (*.exe)")
            if fileName:
                self.msedgeLocationLineEdit.setText(fileName)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while choosing Edge driver location: {e}")

    def chooseSavePathLocation(self):
        try:
            directory = QFileDialog.getExistingDirectory(self, "Select Directory")
            if directory:
                self.savePathLineEdit.setText(directory)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while choosing save path location: {e}")

    def savePrefs(self):
        try:
            self.saveSetting("defaultBrowser", self.browserComboBox.currentText())
            self.saveSetting("savePath", self.savePathLineEdit.text())
            self.saveSetting("driverLocation", self.driverLocationLineEdit.text())
            self.saveSetting("msedgeLocation", self.msedgeLocationLineEdit.text())
            self.saveSetting("proofhubAPIKey", self.proofhubAPIKey.text())
            self.saveSetting("proofhubProjectID", self.proofhubProjectID.text())
            self.saveSetting("proofhubTaskListID", self.proofhubTaskListID.text())
            self.saveSetting("jiraURL", self.jiraURLLineEdit.text())
            self.saveSetting("jiraUsername", self.jiraUserLineEdit.text())
            self.saveSetting("jiraAPIToken", self.jiraTokenLineEdit.text())
            self.saveSetting("mondayURL", self.mondayURLLineEdit.text())
            self.saveSetting("mondayAPIToken", self.mondayTokenLineEdit.text())
            self.saveSetting("clickupAPIToken", self.clickupTokenLineEdit.text())

            self.browserLabel.setText("Browser: " + self.loadSetting("defaultBrowser", "Chrome"))

            self.prefsWindow.close()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving preferences: {e}")

    def settingsFilePath(self):
        return os.path.join(os.getenv('APPDATA'), 'Atom8', 'settings.json')

    def recentFilesFilePath(self):
        return os.path.join(os.getenv('APPDATA'), 'Atom8', 'recent_files.json')

    def saveSetting(self, key, value):
        try:
            settings = self.loadSettings()
            settings[key] = value
            os.makedirs(os.path.dirname(self.settingsFilePath()), exist_ok=True)
            with open(self.settingsFilePath(), 'w') as file:
                json.dump(settings, file)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving setting: {e}")

    def loadSetting(self, key, defaultValue=None):
        settings = self.loadSettings()
        return settings.get(key, defaultValue)

    def loadSettings(self):
        try:
            if os.path.exists(self.settingsFilePath()):
                with open(self.settingsFilePath(), 'r') as file:
                    return json.load(file)
            return {}
        except (FileNotFoundError, json.JSONDecodeError):
            self.logger.warning("Failed to load settings file.")
            return {}

    def loadRecentFiles(self):
        try:
            if os.path.exists(self.recentFilesFilePath()):
                with open(self.recentFilesFilePath(), 'r') as file:
                    self.recentFiles = json.load(file)
                    self.updateRecentFilesMenu()
            else:
                self.recentFiles = []
        except (FileNotFoundError, json.JSONDecodeError):
            self.logger.warning("Failed to load recent files file.")

    def saveRecentFiles(self):
        try:
            os.makedirs(os.path.dirname(self.recentFilesFilePath()), exist_ok=True)
            with open(self.recentFilesFilePath(), 'w') as file:
                json.dump(self.recentFiles, file)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving recent files: {e}")

    def setupScriptEditor(self):
        self.scriptEditorWindow = QMainWindow(self)
        self.scriptEditorWindow.setWindowTitle("AQL Script Editor")
        self.scriptEditorWindow.setGeometry(100, 100, 600, 400)

        self.scriptEditor = QTextEdit(self.scriptEditorWindow)
        self.scriptEditor.setPlaceholderText("Enter AQL script here...")
        self.scriptEditor.setReadOnly(False)
        self.scriptEditorWindow.setCentralWidget(self.scriptEditor)

        self.scriptEditorStatusBar = QStatusBar()
        self.scriptEditorWindow.setStatusBar(self.scriptEditorStatusBar)

        self.scriptEditorMenuBar = self.scriptEditorWindow.menuBar()
        self.scriptEditorFileMenu = self.scriptEditorMenuBar.addMenu('File')

        self.scriptEditorOpenAction = QAction('Open', self)
        self.scriptEditorOpenAction.triggered.connect(self.openScriptFile)

        self.scriptEditorSaveAction = QAction('Save', self)
        self.scriptEditorSaveAction.triggered.connect(self.saveScriptFile)

        self.scriptEditorSaveAsAction = QAction('Save As', self)
        self.scriptEditorSaveAsAction.triggered.connect(self.saveScriptFileAs)

        self.scriptEditorClearAction = QAction('Clear', self)
        self.scriptEditorClearAction.triggered.connect(self.clearScriptEditor)

        self.scriptEditorCloseAction = QAction('Close', self)
        self.scriptEditorCloseAction.triggered.connect(self.closeScriptEditor)

        self.scriptEditorFileMenu.addAction(self.scriptEditorOpenAction)
        self.scriptEditorFileMenu.addAction(self.scriptEditorSaveAction)
        self.scriptEditorFileMenu.addAction(self.scriptEditorSaveAsAction)
        self.scriptEditorFileMenu.addAction(self.scriptEditorClearAction)
        self.scriptEditorFileMenu.addAction(self.scriptEditorCloseAction)

        self.scriptEditorStatusBar.showMessage("Ready", 5000)

        self.scriptEditorLogger = logging.getLogger('ScriptEditor')
        self.scriptEditorLogger.setLevel(logging.INFO)

        self.scriptEditorLoggerTextBox = QTextEditLogger(self.scriptEditor)
        self.scriptEditorLoggerTextBox.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        self.scriptEditorLogger.addHandler(self.scriptEditorLoggerTextBox)
        self.scriptEditorLogger.setLevel(logging.DEBUG)

    def openScriptFile(self):
        fileName, _ = QFileDialog.getOpenFileName(self, "Open File", "", "All Files (*);;AQL Files (*.aql)")
        if fileName:
            try:
                with open(fileName, "r") as file:
                    self.scriptEditor.setText(file.read())
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to open file: {e}")

    def saveScriptFile(self):
        if self.currentFilePath:
            with open(self.currentFilePath, "w") as file:
                file.write(self.scriptEditor.toPlainText())
            self.scriptEditorStatusBar.showMessage("File saved successfully.", 5000)
        else:
            self.saveScriptFileAs()

    def saveScriptFileAs(self):
        fileName, _ = QFileDialog.getSaveFileName(self, "Save As", "", "AQL Files (*.aql)")
        if fileName:
            self.currentFilePath = fileName
            with open(fileName, "w+") as file:
                file.write(self.scriptEditor.toPlainText())
            self.scriptEditorStatusBar.showMessage("File saved as new file.", 5000)

    def clearScriptEditor(self):
        self.scriptEditor.clear()

    def closeScriptEditor(self):
        self.scriptEditorWindow.close()

    def showScriptEditor(self):
        self.scriptEditorWindow.show()

    def newFile(self):
        try:
            if self.steps:
                msgBox = QMessageBox()
                msgBox.setIcon(QMessageBox.Question)
                msgBox.setText("Do you want to save the current file?")
                msgBox.setWindowTitle("Save File")
                msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
                msgBox.setDefaultButton(QMessageBox.Yes)
                response = msgBox.exec()

                if response == QMessageBox.Yes:
                    self.realSaveFile()
                    self.clearStepsList()
                    self.clearInputFields()
                elif response == QMessageBox.No:
                    self.clearStepsList()
                    self.clearInputFields()
                elif response == QMessageBox.Cancel:
                    return
            else:
                self.clearStepsList()
                self.clearInputFields()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while creating new file: {e}")

    def generateBugForJira(self):
        try:
            bugReport = ""
            osName = platform.system().lower()

            testName = self.testName.text()
            testDescription = self.testDescription.text()

            currentDateTime = datetime.now().strftime("%Y-%m-%d %H:%M")
            bugReport += f"**Date:** {currentDateTime}"
            bugReport += f"\n**Name:** {testName}"
            bugReport += f"\n**Description:** {testDescription}\n"

            if osName == "windows":
                bugReport += f"\n**Operating System:** {platform.system()} {platform.release()} {platform.version()}"
            elif osName == "darwin":
                bugReport += f"\n**Operating System:** {platform.system()} {platform.release()} {platform.version()}"

            browserName = self.loadSetting("defaultBrowser", "Chrome")
            browserVersion = ""
            if browserName == "Chrome":
                browserVersion = self.driver.capabilities['browserVersion']
            elif browserName == "Edge":
                browserVersion = self.driver.capabilities['browserVersion']
            elif browserName == "Firefox":
                browserVersion = self.driver.capabilities['browserVersion']
            elif browserName == "Safari":
                browserVersion = self.driver.capabilities['browserVersion']
            bugReport += f"\n**Browser:** {browserName} {browserVersion}\n"
            bugReport += "\n---\n"

            bugReport += "**Performed with the following options:**\n"

            if any(checkbox.isChecked() for checkbox in self.findChildren(QCheckBox)):
                bugReport += "- [x] "
                bugReport += "\n- [x] ".join(
                    [checkbox.text() for checkbox in self.findChildren(QCheckBox) if checkbox.isChecked()])
                bugReport += "\n"
            else:
                bugReport += "- [x] None\n"

            bugReport += "\n---\n"

            bugReport += "**Steps to Reproduce:**\n"
            bugReport += "| Step Number | Action | Value  | Expected Result | Actual Result | Status |\n"
            bugReport += "|-------------|--------|--------|-----------------|---------------|--------|\n"
            for index, step in enumerate(self.steps):
                # STEP
                stepNumber = index + 1
                action = step[0]
                value = ""
                expected = ""
                if action == 'Sleep':
                    value = step[1]
                    expected = f'Sleep for {step[1]} seconds.'
                elif action in ['Click Element', 'Input Text']:
                    value = f'Text: "{step[3]}"' if step[3] else f'Locator: {step[1]}, {step[2]}' if step[2] else ""
                    expected = f'{action} by {step[1]}: {step[2]} {"with text " + step[3] if step[3] else ""}'
                elif action in ['Navigate to URL', 'Execute JavaScript', 'Execute Python Script']:
                    value = step[1]
                    expected = f'{action}: {step[1]}'
                elif action == 'Take Screenshot':
                    value = step[1]
                    expected = f'Take screenshot: {step[1]}'
                elif action == 'Maximize Window':
                    expected = 'Maximize Window'
                elif action == 'Compare Images':
                    value = f'Image 1: {step[1]}, Image 2: {step[2]}'
                    expected = f'Compare images: {step[1]} and {step[2]}'
                else:
                    expected = 'Unknown Action'

                if self.results[index][1] == 'Passed':
                    bugReport += f"| {stepNumber} | {action} | {value} | {expected} | | Passed |\n"
                else:
                    bugReport += f"| {stepNumber} | {action} | {value} | {expected} | | Failed |\n"

            bugReport += "\n"
            bugReport += "\n**Attachments:**\n <Add screenshots here>\n"
            bugReport += "\n---\n"
            return bugReport
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate bug report: {e}")

    def updateRecentFiles(self, fileName):
        try:
            if fileName in self.recentFiles:
                self.recentFiles.remove(fileName)
            self.recentFiles.insert(0, fileName)
            if len(self.recentFiles) > 10:
                self.recentFiles.pop()
            self.saveRecentFiles()
            self.updateRecentFilesMenu()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating recent files: {e}")

    def saveLogs(self):
        try:
            options = QFileDialog.Options()
            fileName, _ = QFileDialog.getSaveFileName(self, "Save Logs", "", "Log Files (*.log)", options=options)
            if fileName:
                with open(fileName, "w") as file:
                    file.write(self.logViewer.toPlainText())
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving logs: {e}")

    def clearLogs(self):
        self.logViewer.clear()

    def editSelectedStep(self):
        # STEP
        try:
            selected_item = self.stepsList.currentRow()
            if selected_item >= 0:
                self.editMode = True
                self.editIndex = selected_item
                self.editButton.setVisible(False)
                self.saveButton.setVisible(True)

                step = self.steps[selected_item]
                action = step[0]
                self.actionSelection.setCurrentText(action)
                if action in ['Click Element', 'Input Text']:
                    self.locatorSelection.setVisible(True)
                    self.locatorInput.setVisible(True)
                    self.locatorSelection.setCurrentText(step[1])
                    self.locatorInput.setText(step[2])
                else:
                    self.locatorSelection.setVisible(False)
                    self.locatorInput.setVisible(False)

                self.inputText.setVisible(
                    action in ['Input Text', 'Execute Python Script', 'Execute JavaScript', 'Navigate to URL',
                               'Take Screenshot'])
                self.sleepInput.setVisible(action == 'Sleep')
                self.inputDescription.setVisible(
                    action in ['Click Element', 'Input Text', 'Sleep', 'Navigate to URL', 'Execute Python Script',
                               'Execute JavaScript'])

                if action == 'Navigate to URL':
                    self.inputText.setPlaceholderText("Enter URL")
                    self.inputText.setText(step[1])
                elif action == 'Input Text':
                    self.inputText.setPlaceholderText("Enter Text")
                    self.inputText.setText(step[3])
                elif action == 'Execute Python Script':
                    self.inputText.setPlaceholderText("Enter Script Path")
                    self.inputText.setText(step[1])
                elif action == 'Execute JavaScript':
                    self.inputText.setPlaceholderText("Enter JavaScript Code")
                    self.inputText.setText(step[1])
                elif action == 'Sleep':
                    self.sleepInput.setPlaceholderText("Enter Sleep Time")
                    self.sleepInput.setText(step[1])
                elif action == 'Take Screenshot':
                    self.inputText.setPlaceholderText("Enter Screenshot Name")
                    self.inputText.setText(step[1])
                elif action == 'Compare Images':
                    self.refImgPath.setVisible(True)
                    self.refImgPath.setText(step[1])
                    self.compImgPath.setVisible(True)
                    self.compImgPath.setText(step[2])
                    self.inputText.setVisible(False)
                    self.sleepInput.setVisible(False)
                    self.locatorSelection.setVisible(False)
                    self.locatorInput.setVisible(False)
                else:
                    self.inputText.setPlaceholderText("Enter Text")
                    self.inputText.setText(step[1])

                self.inputDescription.setText(step[-1])
            else:
                QMessageBox.warning(self, "No Selection", "Please select a step to edit.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while editing step: {e}")

    def updateStep(self):
        # STEP
        try:
            selected_item = self.stepsList.currentRow()
            if selected_item >= 0:
                action = self.actionSelection.currentText()
                locator_type = self.locatorSelection.currentText()
                locator_value = self.locatorInput.text()
                text_value = self.inputText.text()
                sleep_value = self.sleepInput.text()
                description_value = self.inputDescription.text()

                if action == 'Sleep':
                    step = (action, sleep_value)
                    display_txt = f'Sleep for {sleep_value} seconds.'
                elif action in ['Click Element', 'Input Text']:
                    step = (action, locator_type, locator_value, text_value, description_value)
                    display_txt = f'{action}: {locator_value if locator_value else text_value} (Locator: {locator_type if locator_type != "Select Locator" else "N/A"}), Description: {description_value}'
                elif action in ['Navigate to URL', 'Execute JavaScript', 'Execute Python Script']:
                    step = (action, text_value, description_value)
                    display_txt = f'{action}: {text_value}{"." if not description_value else f", Description: {description_value}"}'
                elif action == 'Take Screenshot':
                    step = (action, text_value, description_value)
                    display_txt = f'Take screenshot and save as {text_value}{".png" if not text_value.endswith(".png") else ""}{"." if not description_value else f", Description: {description_value}"}'
                elif action == 'Maximize Window':
                    step = action
                    display_txt = f'Maximize Window.'
                elif action == 'Compare Images':
                    step = (action, self.refImgPath.text(), self.compImgPath.text())
                    display_txt = f'Comparing images: {self.refImgPath.text()} and {self.compImgPath.text()}'
                else:
                    QMessageBox.warning(self, "Invalid Action", "The selected action is not supported.")
                    return

                self.steps[selected_item] = step
                self.stepsList.item(selected_item).setText(display_txt)
            else:
                QMessageBox.warning(self, "No Selection", "Please select a step to update.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating step: {e}")

    def moveStepUp(self):
        try:
            selected_item = self.stepsList.currentRow()
            if selected_item >= 1:
                self.stepsList.insertItem(selected_item - 1, self.stepsList.takeItem(selected_item))
                self.steps.insert(selected_item - 1, self.steps.pop(selected_item))
                self.stepsList.setCurrentRow(selected_item - 1)
            else:
                QMessageBox.warning(self, "No Selection", "Please select a step to move up.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while moving step up: {e}")

    def moveStepDown(self):
        try:
            selected_item = self.stepsList.currentRow()
            if selected_item >= 0 and selected_item < self.stepsList.count() - 1:
                self.stepsList.insertItem(selected_item + 1, self.stepsList.takeItem(selected_item))
                self.steps.insert(selected_item + 1, self.steps.pop(selected_item))
                self.stepsList.setCurrentRow(selected_item + 1)
            else:
                QMessageBox.warning(self, "No Selection", "Please select a step to move down.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while moving step down: {e}")

    def addOrEditStep(self):
        try:
            if self.editMode:
                self.updateStep()
                self.editMode = False
                self.editIndex = None
                self.editButton.setVisible(True)
                self.saveButton.setVisible(False)
            else:
                self.addStep()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while adding or editing step: {e}")

    def saveEditedStep(self):
        try:
            self.updateStep()
            self.editMode = False
            self.editIndex = None
            self.editButton.setVisible(True)
            self.saveButton.setVisible(False)
            self.clearInputFields()

            self.logger.info("Saved edited step.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving edited step: {e}")

    def clearInputFields(self):
        try:
            self.actionSelection.setCurrentIndex(0)
            self.locatorSelection.setCurrentIndex(0)
            self.locatorInput.setText('')
            self.inputText.setText('')
            self.sleepInput.setText('')
            self.inputDescription.setText('')
            self.refImgPath.setText('')
            self.compImgPath.setText('')
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while clearing input fields: {e}")

    def updateRecentFilesMenu(self):
        try:
            self.recentFilesMenu.clear()
            for index, filePath in enumerate(self.recentFiles):
                action = QAction(f"{index + 1}. {filePath.split('/')[-1]}", self)
                action.triggered.connect(lambda checked, path=filePath: self.openRecentFile(path))
                self.recentFilesMenu.addAction(action)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating recent files menu: {e}")

    def openRecentFile(self, filePath):
        try:
            with open(filePath, "r") as file:
                file_content = json.load(file)
                if not isinstance(file_content, dict) or "steps" not in file_content:
                    raise ValueError("File content is not in the expected format")

                self.testName.setText(file_content.get("testName", ""))
                self.testDescription.setText(file_content.get("testDescription", ""))
                self.steps = file_content["steps"]
                self.stepsList.clear()
                for index, step in enumerate(self.steps):
                    if not isinstance(step, (list, tuple)):
                        QMessageBox.critical(self, "Error", f"Invalid step format at index {index}: {step}")
                        continue
                    display_text = self.constructStepDisplayText(step)
                    self.stepsList.addItem(display_text)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open file: {e}")

    def loadRecentFiles(self):
        try:
            if os.path.exists(self.recentFilesFilePath()):
                with open(self.recentFilesFilePath(), 'r') as file:
                    self.recentFiles = json.load(file)
                    self.updateRecentFilesMenu()
            else:
                self.recentFiles = []
        except (FileNotFoundError, json.JSONDecodeError):
            self.logger.warning("Failed to load recent files file.")

    def extractWebElements(self):
        try:
            url, ok = QInputDialog.getText(self, 'Extract Web Elements', 'Enter the URL:')
            if ok and url:
                try:
                    elements_data = extract_elements_to_json(url)
                    self.showExtractionResult(elements_data)
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to extract elements: {e}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while extracting web elements: {e}")

    def showSequencer(self):
        try:
            self.sequencerWindow = QDialog(self, Qt.Window)
            self.sequencerWindow.setWindowTitle("Sequencer")
            sequencerLayout = QVBoxLayout()

            menuBar = QMenuBar()
            fileMenu = menuBar.addMenu("Options")

            createSequenceAction = QAction("Create Sequence", self)
            createSequenceAction.triggered.connect(self.chooseAtm8File)
            fileMenu.addAction(createSequenceAction)

            saveSequenceAction = QAction("Save Sequence", self)
            saveSequenceAction.triggered.connect(self.saveSequence)
            fileMenu.addAction(saveSequenceAction)

            loadSequenceAction = QAction("Load Sequence", self)
            loadSequenceAction.triggered.connect(self.loadSequence)
            fileMenu.addAction(loadSequenceAction)

            fileMenu.addSeparator()

            runSequenceAction = QAction("Run Sequence", self)
            runSequenceAction.triggered.connect(self.runSequencer)
            fileMenu.addAction(runSequenceAction)

            sequencerLayout.setMenuBar(menuBar)

            buttonLayout = QHBoxLayout()

            self.chooseFileButton = QPushButton("Choose Files")
            self.chooseFileButton.clicked.connect(self.chooseAtm8File)
            buttonLayout.addWidget(self.chooseFileButton)

            self.removeFileButton = QPushButton("Remove File")
            self.removeFileButton.clicked.connect(self.removeAtm8File)
            self.removeFileButton.setProperty("class", "secondary-btn")
            buttonLayout.addWidget(self.removeFileButton)

            self.loadToMainEditorButton = QPushButton("Load to Main Editor")
            self.loadToMainEditorButton.clicked.connect(self.loadToMainEditor)
            self.loadToMainEditorButton.setProperty("class", "secondary-btn")
            buttonLayout.addWidget(self.loadToMainEditorButton)

            self.runSequencerButton = QPushButton("Run Sequence")
            self.runSequencerButton.clicked.connect(self.runSequencer)
            self.runSequencerButton.setProperty("class", "primary-btn")
            buttonLayout.addWidget(self.runSequencerButton)

            self.runSequencerButton.setStyleSheet("""
                QPushButton {
                    color: white;
                    background-color: #28A745;
                    border-radius: 4px;
                    padding: 6px 12px;
                    border: none;
                    font-size: 12px;
                }

                QPushButton:hover {
                    background-color: #218838;
                    border-color: #1E7E34;
                }

                QPushButton:pressed {
                    background-color: #1D7D33;
                    border-color: #1C7430;
                }
            """)

            self.generateReport = QCheckBox("Generate Report")
            sequencerLayout.addWidget(self.generateReport)

            sequencerLayout.addLayout(buttonLayout)

            self.atm8FilesList = QListWidget()
            sequencerLayout.addWidget(self.atm8FilesList)

            self.sequencerWindow.setLayout(sequencerLayout)
            self.sequencerWindow.resize(600, 400)
            self.sequencerWindow.show()


        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing sequencer: {e}")

    def chooseAtm8File(self):
        try:
            options = QFileDialog.Options()
            fileNames, _ = QFileDialog.getOpenFileNames(self, "Choose Files", "", "Atom8 Files (*.atm8)",
                                                        options=options)
            if fileNames:
                self.atm8FilesList.addItems(fileNames)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while choosing atm8 file: {e}")

    def runSequencer(self, fileNames):
        try:
            self.steps.clear()
            for i in range(self.atm8FilesList.count()):
                file_name = self.atm8FilesList.item(i).text()
                with open(file_name, "r") as file:
                    file_content = json.load(file)
                    if not isinstance(file_content, dict) or "steps" not in file_content:
                        raise ValueError("File content is not in the expected format")

                    for step in file_content["steps"]:
                        self.steps.append(step)
                        display_text = self.constructStepDisplayText(step)
                        self.stepsList.addItem(display_text)
            self.logger.info("Running sequencer.")
            self.startAutomation()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while running sequencer: {e}")

    def removeAtm8File(self):
        try:
            selected_item = self.atm8FilesList.currentRow()
            if selected_item >= 0:
                self.atm8FilesList.takeItem(selected_item)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while removing atm8 file: {e}")

    def loadToMainEditor(self):
        try:
            selected_item = self.atm8FilesList.currentRow()
            if selected_item >= 0:
                file_name = self.atm8FilesList.item(selected_item).text()
                self.openRecentFile(file_name)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while loading to main editor: {e}")

    def loadSequence(self):
        try:
            options = QFileDialog.Options()
            fileName, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Sequence Files (*.seq)", options=options)
            if fileName:
                with open(fileName, "r") as file:
                    sequence_content = json.load(file)
                    if not isinstance(sequence_content, list):
                        raise ValueError("Sequence content is not in the expected format")

                    for file_name in sequence_content:
                        self.atm8FilesList.addItem(file_name)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open file: {e}")

    def saveSequence(self):
        try:
            options = QFileDialog.Options()
            fileName, _ = QFileDialog.getSaveFileName(self, "Save Sequence", "", "Sequence Files (*.seq)",
                                                      options=options)
            if fileName:
                sequence_content = [self.atm8FilesList.item(i).text() for i in range(self.atm8FilesList.count())]
                with open(fileName, "w") as file:
                    json.dump(sequence_content, file)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while saving sequence: {e}")

    def compareImages(self, reference_path, test_path, output_path):
        self.driver.save_screenshot(test_path)

        reference = cv2.imread(reference_path)
        with open(self.settingsFilePath(), "r") as settings_file:
            settings = json.load(settings_file)
            screenshot_folder = settings.get("savePath")
            if not os.path.isdir(screenshot_folder):
                os.makedirs(screenshot_folder)
            screenshot_filename = os.path.join(screenshot_folder, test_path)
            self.driver.save_screenshot(screenshot_filename)
            self.logger.info(f"[Compare Images] -> Screenshot saved as {screenshot_filename}")

            test = cv2.imread(screenshot_filename)

        difference = cv2.absdiff(reference, test)
        gray = cv2.cvtColor(difference, cv2.COLOR_BGR2GRAY)

        _, thresh = cv2.threshold(gray, 1, 255, cv2.THRESH_BINARY)
        contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        overlay = test.copy()
        overlay[:] = (0, 0, 0)

        cv2.addWeighted(overlay, 0.5, test, 0.5, 0, test)

        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            cv2.rectangle(test, (x, y), (x + w, y + h), (0, 255, 0), 1)

        cv2.imwrite(output_path, test)

        return output_path

    def showExtractionResult(self, elements_data):
        try:
            self.resultsWindow = QDialog(self, Qt.Window)
            self.resultsWindow.setWindowTitle("Extraction Results")
            resultsLayout = QVBoxLayout()

            self.urlInputField = QLineEdit(self.resultsWindow)
            self.urlInputField.setPlaceholderText('Enter URL')
            resultsLayout.addWidget(self.urlInputField)

            searchButton = QPushButton('Extract Web Elements', self.resultsWindow)
            searchButton.clicked.connect(self.onSearchClicked)
            resultsLayout.addWidget(searchButton)

            self.resultsTable = QTableWidget()
            self.updateResultsTable(elements_data)
            self.resultsTable.setContextMenuPolicy(Qt.CustomContextMenu)
            self.resultsTable.customContextMenuRequested.connect(self.resultsTableContextMenu)
            resultsLayout.addWidget(self.resultsTable)

            self.resultsWindow.setLayout(resultsLayout)
            self.resultsWindow.resize(600, 400)
            self.resultsWindow.show()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing extraction results: {e}")

    def resultsTableContextMenu(self, position):
        try:
            index = self.resultsTable.indexAt(position)
            if index.isValid() and index.column() == 1:
                menu = QMenu()
                copyAction = menu.addAction("Copy")
                action = menu.exec_(self.resultsTable.viewport().mapToGlobal(position))
                if action == copyAction:
                    self.copyLocatorValue(index.row())
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while showing results table context menu: {e}")

    def copyLocatorValue(self, row):
        try:
            comboBox = self.resultsTable.cellWidget(row, 1)
            if comboBox:
                selectedText = comboBox.currentText()
                clipboard = QApplication.clipboard()
                clipboard.setText(selectedText)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while copying locator value: {e}")

    def updateResultsTable(self, elements_data):
        try:
            self.resultsTable.setColumnCount(2)
            self.resultsTable.setHorizontalHeaderLabels(['Value', 'Locators'])
            self.resultsTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.resultsTable.verticalHeader().setVisible(False)
            self.resultsTable.setEditTriggers(QTableWidget.NoEditTriggers)

            self.resultsTable.setRowCount(len(elements_data))
            for row, element in enumerate(elements_data):
                self.resultsTable.setRowHeight(row, 45)
                self.resultsTable.setItem(row, 0, QTableWidgetItem(element['value']))
                comboBox = CustomComboBox()
                for locator in element['locators']:
                    comboBox.addItem(f"XPath: {locator['xpath']}")
                    for attr, value in locator['attributes'].items():
                        comboBox.addItem(f"{attr}: {value}")
                self.resultsTable.setCellWidget(row, 1, comboBox)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while updating results table: {e}")

    def onSearchClicked(self):
        try:
            url = self.urlInputField.text()
            if url:
                QApplication.processEvents()

                try:
                    elements_data = extract_elements_to_json(url)
                    self.updateResultsTable(elements_data)
                    self.statusBar.clearMessage()
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to extract elements: {e}")
                    self.statusBar.clearMessage()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error while searching: {e}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Atom8()
    ex.show()
    sys.exit(app.exec_())
